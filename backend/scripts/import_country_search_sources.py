from __future__ import annotations

import argparse
import asyncio
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from sqlalchemy import delete

from app.database import SessionLocal
from app.models.country_search_source import CountrySearchSource
from app.services.search.country_sources import load_country_search_source_cache


SHEET_CONFIG = {
    "YellowPages_Top3": "directory",
    "B2B_Top3": "marketplace",
}


def _normalized_domain(url: str | None) -> str | None:
    host = (urlparse(str(url or "").strip()).netloc or "").strip().lower()
    if not host:
        return None
    return host.removeprefix("www.")


def _iter_source_rows(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []
    for sheet_name, source_type in SHEET_CONFIG.items():
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=5, values_only=True):
            country_name, country_code, continent = row[0], row[1], row[2]
            selection_mode, method_note = row[9], row[10]
            if not country_code:
                continue
            for rank, (source_name, source_url) in enumerate(((row[3], row[4]), (row[5], row[6]), (row[7], row[8])), start=1):
                if not source_name or not source_url:
                    continue
                domain = _normalized_domain(str(source_url))
                if not domain:
                    continue
                rows.append(
                    {
                        "country_code": str(country_code).strip().upper(),
                        "country_name": str(country_name or "").strip(),
                        "continent": str(continent or "").strip() or None,
                        "source_type": source_type,
                        "source_rank": rank,
                        "source_name": str(source_name).strip(),
                        "source_url": str(source_url).strip(),
                        "source_domain": domain,
                        "selection_mode": str(selection_mode or "").strip() or None,
                        "method_note": str(method_note or "").strip() or None,
                    }
                )
    return rows


async def _import_rows(workbook_path: Path) -> int:
    rows = _iter_source_rows(workbook_path)
    source_models = [CountrySearchSource(**row) for row in rows]
    async with SessionLocal() as session:
        await session.execute(delete(CountrySearchSource))
        session.add_all(source_models)
        await session.commit()
    await load_country_search_source_cache()
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import country yellow pages and B2B sources from workbook.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default="/Users/Maxwell/Desktop/global_yellowpages_b2b_top3_249_markets.xlsx",
        help="Path to the source workbook.",
    )
    args = parser.parse_args()
    count = asyncio.run(_import_rows(Path(args.workbook).expanduser()))
    print(f"imported_rows\t{count}")


if __name__ == "__main__":
    main()
