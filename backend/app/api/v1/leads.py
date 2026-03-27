from __future__ import annotations

import asyncio
import csv
import math
from datetime import datetime, timezone
from io import BytesIO, StringIO
from urllib.parse import urlparse
from uuid import UUID

import httpx
from bs4 import BeautifulSoup
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import delete, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.database import SessionLocal, get_db
from app.models.company import Company
from app.models.lead import Lead
from app.models.lead_review import LeadReview
from app.models.search_keyword import SearchKeyword
from app.models.search_keyword_company import SearchKeywordCompany
from app.models.task import Task
from app.schemas.lead import LeadListResponse, LeadSearchRequest
from app.schemas.review import LeadReviewAnnotation, LeadReviewUpsertRequest, REVIEWABLE_FIELD_KEYS
from app.schemas.task import TaskCreateResponse
from app.services.extraction.company_name import CompanyNameExtractor
from app.services.extraction.country_detection import CountryDetector, resolve_country
from app.services.extraction.relevance import IndustryRelevanceClassifier
from app.services.extraction.social_links import SocialLinksExtractor
from app.services.search.keyword_cache import (
    ALL_QUERY_TEMPLATES,
    PRIMARY_QUERY_TEMPLATES,
    SECONDARY_QUERY_TEMPLATES,
    build_keyword_queries,
    canonical_domain,
    next_refresh_at,
    normalize_keywords,
    scope_fingerprint,
)
from app.services.search.serper import SerperClient
from app.services.workspace_store import ROOT_TASK_TYPE, cleanup_old_root_tasks, get_root_task, get_task_leads

router = APIRouter(prefix="/leads", tags=["leads"])

SETTINGS = get_settings()
KEYWORD_TASK_TYPE = "lead_search_keyword"
IGNORED_HOSTS = {
    "example.com",
    "example.org",
    "example.net",
    "linkedin.com",
    "www.linkedin.com",
    "facebook.com",
    "www.facebook.com",
}
DEMO_SUFFIXES = [
    "Trading",
    "Manufacturing",
    "Supply",
    "Industries",
    "Global",
    "Solutions",
    "Exports",
    "Partners",
]
AVG_SERPER_PAGE_SECONDS = 1.5
AVG_CANDIDATE_BUILD_SECONDS = 5.0
FINALIZE_SECONDS = 2
DEMO_TOTAL_SECONDS = 2
QUERY_PAGE_SIZE = 10
LEADS: dict[str, list] = {}
CONTACTS: dict[str, list] = {}


class SearchRuntime:
    def __init__(self) -> None:
        self.settings = SETTINGS
        self.serper_client = SerperClient()
        self.relevance_classifier = IndustryRelevanceClassifier()
        self.search_semaphore = asyncio.Semaphore(max(1, self.settings.max_concurrent_searches))
        self.scrape_semaphore = asyncio.Semaphore(max(1, self.settings.max_concurrent_scrapers))


async def _safe_serper_search(serper_client: SerperClient, query: str, hl: str, gl: str, num: int, page: int = 1) -> dict:
    try:
        return await serper_client.search(query=query, hl=hl, gl=gl, num=num, page=page)
    except Exception:
        return {"organic": []}


async def _fetch_homepage(website: str, timeout: float = 15.0) -> tuple[BeautifulSoup | None, str | None]:
    try:
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            response = await client.get(website, headers={"User-Agent": "Mozilla/5.0 LeadGenBot/1.0"})
            response.raise_for_status()
    except Exception:
        return None, None
    return BeautifulSoup(response.text, "html.parser"), response.text


def _keywords_from_params(params: dict | None) -> list[str]:
    if not params:
        return []
    return normalize_keywords(list(params.get("keywords") or []) + ([params["product_name"]] if params.get("product_name") else []))


def _field_source(source_type: str, source_url: str | None, extractor: str, source_hint: str | None = None) -> dict:
    return {
        "source_type": source_type,
        "source_url": source_url,
        "extractor": extractor,
        "source_hint": source_hint,
    }


def _merge_field_provenance(base: dict | None, updates: dict[str, dict | None]) -> dict:
    merged = dict(base or {})
    for field_key, value in updates.items():
        if value:
            merged[field_key] = value
    return merged


def _review_annotation(review: LeadReview) -> LeadReviewAnnotation:
    return LeadReviewAnnotation(
        verdict=review.verdict,
        source_path=review.source_path,
        note=review.note,
        updated_at=review.updated_at,
    )


def _build_queries(
    payload: LeadSearchRequest,
    keyword: str | None = None,
    stage: int = 1,
) -> list[dict[str, str | int]]:
    search_keyword = keyword or payload.keywords[0]
    return build_keyword_queries(
        search_keyword,
        countries=payload.countries,
        languages=payload.languages,
        stage=stage,
    )


def _keyword_limit(payload: LeadSearchRequest) -> int | None:
    return payload.target_count


def _max_pages(target_count: int | None) -> int:
    if target_count is None:
        return 3
    return max(1, min(10, math.ceil(target_count / 50)))


def _pages_per_query(payload: LeadSearchRequest) -> int:
    pages = _max_pages(payload.target_count)
    if payload.mode == "live" and payload.countries:
        return max(2, pages)
    return pages


def _slugify(value: str) -> str:
    slug = "".join(character.lower() if character.isalnum() else " " for character in value)
    return "-".join(part for part in slug.split() if part) or "keyword"


def _demo_result_count(payload: LeadSearchRequest) -> int:
    if payload.target_count is not None:
        return payload.target_count
    country_count = len(payload.countries) if payload.countries else 1
    language_count = len(payload.languages) if payload.languages else 1
    keyword_count = len(payload.keywords) if payload.keywords else 1
    return max(3, min(12, country_count * language_count * keyword_count))


def _initial_candidate_budget(target_count: int | None, query_count: int, pages_per_query: int) -> int:
    if target_count is None:
        return min(query_count * pages_per_query * 10, query_count * 30)
    return max(target_count, target_count * 2)


def _estimate_keyword_runtime(payload: LeadSearchRequest, keyword: str, candidate_budget_override: int | None = None) -> dict[str, int]:
    if payload.mode == "demo":
        planned_candidate_budget = candidate_budget_override or _demo_result_count(payload)
        return {
            "planned_search_requests": 0,
            "planned_candidate_budget": planned_candidate_budget,
            "estimated_total_seconds": DEMO_TOTAL_SECONDS,
        }
    query_count = len(_build_queries(payload, keyword=keyword, stage=1)) + len(_build_queries(payload, keyword=keyword, stage=2))
    pages_per_query = _pages_per_query(payload)
    planned_search_requests = query_count * pages_per_query
    planned_candidate_budget = candidate_budget_override or _initial_candidate_budget(payload.target_count, max(1, query_count), pages_per_query)
    search_seconds = math.ceil(planned_search_requests / max(1, SETTINGS.max_concurrent_searches)) * AVG_SERPER_PAGE_SECONDS
    candidate_seconds = planned_candidate_budget * AVG_CANDIDATE_BUILD_SECONDS
    estimated_total_seconds = max(1, int(math.ceil(search_seconds + candidate_seconds + FINALIZE_SECONDS)))
    return {
        "planned_search_requests": planned_search_requests,
        "planned_candidate_budget": planned_candidate_budget,
        "estimated_total_seconds": estimated_total_seconds,
    }


def _estimate_search_runtime(payload: LeadSearchRequest) -> dict[str, int]:
    if payload.mode == "demo":
        budget = _demo_result_count(payload)
        return {
            "planned_search_requests": 0,
            "planned_candidate_budget": budget * max(1, len(payload.keywords)),
            "estimated_total_seconds": DEMO_TOTAL_SECONDS,
        }
    total_search_requests = 0
    total_candidate_budget = 0
    total_estimated_seconds = 0
    for keyword in payload.keywords:
        plan = _estimate_keyword_runtime(payload, keyword)
        total_search_requests += plan["planned_search_requests"]
        total_candidate_budget += plan["planned_candidate_budget"]
        total_estimated_seconds += plan["estimated_total_seconds"]
    return {
        "planned_search_requests": total_search_requests,
        "planned_candidate_budget": total_candidate_budget,
        "estimated_total_seconds": max(1, total_estimated_seconds),
    }


def _build_demo_lead_item(payload: LeadSearchRequest, keyword: str, index: int) -> dict:
    countries = payload.countries or ["Global Market"]
    languages = payload.languages or ["en"]
    country = countries[index % len(countries)]
    language = languages[index % len(languages)]
    suffix = DEMO_SUFFIXES[index % len(DEMO_SUFFIXES)]
    keyword_slug = _slugify(keyword)
    country_slug = _slugify(country)
    company_name = f"{keyword.strip()} {suffix} {index + 1}".strip()
    website = f"https://{keyword_slug}-{country_slug}-{index + 1}.example.com"
    target_country = resolve_country(country)
    return {
        "company_name": company_name,
        "website": website,
        "facebook_url": f"https://www.facebook.com/{keyword_slug}-{index + 1}",
        "linkedin_url": f"https://www.linkedin.com/company/{keyword_slug}-{country_slug}-{index + 1}",
        "country": target_country.name_en if target_country is not None else country,
        "continent": target_country.continent if target_country is not None else "Global",
        "source": "demo",
        "contact_status": "pending",
        "decision_maker_status": "pending",
        "general_contact_status": "pending",
        "contact_name": None,
        "contact_title": None,
        "linkedin_personal_url": None,
        "personal_email": None,
        "work_email": None,
        "phone": None,
        "whatsapp": None,
        "potential_contacts": None,
        "general_emails": [],
        "matched_keywords": [keyword],
        "raw_data": {
            "search_title": company_name,
            "search_snippet": f"Demo lead generated for {keyword} in {country}.",
            "name_source": "demo_mode",
            "demo_mode": True,
            "demo_index": index + 1,
            "demo_language": language,
            "target_country": target_country.name_en if target_country is not None else country,
            "target_country_code": target_country.code if target_country is not None else None,
            "country_detection": {
                "status": "matched",
                "target_country_code": target_country.code if target_country is not None else None,
                "target_country_name": target_country.name_en if target_country is not None else country,
                "detected_country_code": target_country.code if target_country is not None else None,
                "detected_country_name": target_country.name_en if target_country is not None else country,
                "continent": target_country.continent if target_country is not None else "Global",
                "confidence": 1.0,
                "evidence": [{"signal": "demo_mode", "value": country, "weight": 100}],
                "mismatch_reason": None,
            },
            "matched_keywords": [keyword],
        },
    }


def _build_demo_keyword_items(payload: LeadSearchRequest, keyword: str) -> list[dict]:
    return [_build_demo_lead_item(payload, keyword, index) for index in range(_demo_result_count(payload))]


def _task_payload(task: Task) -> LeadSearchRequest:
    params = task.params or {}
    return LeadSearchRequest(
        keywords=_keywords_from_params(params),
        product_name=params.get("product_name"),
        continents=params.get("continents") or [],
        countries=params.get("countries") or [],
        languages=params.get("languages") or [],
        target_count=task.target_count,
        mode=params.get("mode", "live"),
    )


def _refresh_search_task_projection(task: Task) -> None:
    estimated_total = max(1, task.estimated_total_seconds or 1)
    if task.status in {"completed", "failed"}:
        task.progress = 100 if task.status == "completed" else task.progress
        task.estimated_remaining_seconds = 0
        return
    if task.planned_search_requests == 0 and task.planned_candidate_budget == 0:
        task.estimated_remaining_seconds = DEMO_TOTAL_SECONDS
        task.progress = 0
        return
    completed_estimated_seconds = (
        (task.processed_search_requests / max(1, SETTINGS.max_concurrent_searches)) * AVG_SERPER_PAGE_SECONDS
        + task.processed_candidates * AVG_CANDIDATE_BUILD_SECONDS
    )
    task.estimated_remaining_seconds = max(0, int(math.ceil(estimated_total - completed_estimated_seconds)))
    task.progress = min(99, int(round((min(completed_estimated_seconds, estimated_total) / estimated_total) * 100)))
    task.completed = task.processed_candidates


async def _recompute_root_task(root_task_id: UUID) -> None:
    async with SessionLocal() as session:
        root = await session.get(Task, root_task_id)
        if root is None:
            return
        result = await session.execute(
            select(Task).where(Task.parent_task_id == root_task_id, Task.type == KEYWORD_TASK_TYPE)
        )
        children = result.scalars().all()
        root.processed_search_requests = sum(child.processed_search_requests for child in children)
        root.planned_search_requests = sum(child.planned_search_requests for child in children)
        root.processed_candidates = sum(child.processed_candidates for child in children)
        root.planned_candidate_budget = sum(child.planned_candidate_budget for child in children)
        root.completed = root.processed_candidates
        root.total = max(root.planned_candidate_budget, 0)
        root.confirmed_leads = sum(child.confirmed_leads for child in children)
        root.estimated_total_seconds = max(1, sum((child.estimated_total_seconds or 0) for child in children) or 1)
        root.estimated_remaining_seconds = sum((child.estimated_remaining_seconds or 0) for child in children if child.status not in {"completed", "failed"})
        _refresh_search_task_projection(root)
        await session.commit()


async def _mutate_task(
    task_id: UUID,
    *,
    phase: str | None = None,
    status: str | None = None,
    processed_search_requests_inc: int = 0,
    processed_candidates_inc: int = 0,
    planned_candidate_budget: int | None = None,
    confirmed_leads: int | None = None,
    stopped_early: bool | None = None,
    error: str | None = None,
    total: int | None = None,
    progress: int | None = None,
    estimated_remaining_seconds: int | None = None,
    params: dict | None = None,
) -> None:
    root_task_id: UUID | None = None
    async with SessionLocal() as session:
        result = await session.execute(select(Task).where(Task.id == task_id).with_for_update())
        task = result.scalar_one_or_none()
        if task is None:
            return
        if phase is not None:
            task.phase = phase
        if status is not None:
            task.status = status
        if processed_search_requests_inc:
            task.processed_search_requests += processed_search_requests_inc
        if processed_candidates_inc:
            task.processed_candidates += processed_candidates_inc
        if planned_candidate_budget is not None:
            task.planned_candidate_budget = planned_candidate_budget
        if confirmed_leads is not None:
            task.confirmed_leads = confirmed_leads
        if stopped_early is not None:
            task.stopped_early = stopped_early
        if error is not None:
            task.error = error
        if total is not None:
            task.total = total
        if progress is not None:
            task.progress = progress
        if estimated_remaining_seconds is not None:
            task.estimated_remaining_seconds = estimated_remaining_seconds
        if params is not None:
            task.params = params
        task.updated_at = datetime.now(timezone.utc)
        _refresh_search_task_projection(task)
        root_task_id = task.parent_task_id
        await session.commit()
    if root_task_id is not None:
        await _recompute_root_task(root_task_id)


async def _get_or_create_search_keyword(keyword: str, payload: LeadSearchRequest) -> SearchKeyword:
    fingerprint = scope_fingerprint(countries=payload.countries, languages=payload.languages)
    keyword_normalized = normalize_keywords([keyword])[0].lower()
    now = datetime.now(timezone.utc)
    async with SessionLocal() as session:
        result = await session.execute(
            select(SearchKeyword).where(
                SearchKeyword.keyword_normalized == keyword_normalized,
                SearchKeyword.scope_fingerprint == fingerprint,
            )
        )
        row = result.scalar_one_or_none()
        if row is not None:
            row.last_requested_at = now
            await session.commit()
            await session.refresh(row)
            return row
        row = SearchKeyword(
            keyword=keyword,
            keyword_normalized=keyword_normalized,
            scope_fingerprint=fingerprint,
            countries=payload.countries,
            languages=payload.languages,
            query_frontier={},
            refresh_status="pending",
            refresh_error=None,
            last_requested_at=now,
            next_refresh_at=next_refresh_at(now),
        )
        session.add(row)
        try:
            await session.commit()
        except IntegrityError:
            await session.rollback()
            result = await session.execute(
                select(SearchKeyword).where(
                    SearchKeyword.keyword_normalized == keyword_normalized,
                    SearchKeyword.scope_fingerprint == fingerprint,
                )
            )
            row = result.scalar_one()
            row.last_requested_at = now
            await session.commit()
        await session.refresh(row)
        return row


async def _claim_keyword_refresh(search_keyword_id: UUID) -> bool:
    async with SessionLocal() as session:
        result = await session.execute(
            select(SearchKeyword).where(SearchKeyword.id == search_keyword_id).with_for_update()
        )
        keyword_row = result.scalar_one_or_none()
        if keyword_row is None:
            return False
        if keyword_row.refresh_status == "running":
            return False
        keyword_row.refresh_status = "running"
        keyword_row.refresh_error = None
        await session.commit()
        return True


async def _complete_keyword_refresh(search_keyword_id: UUID) -> None:
    async with SessionLocal() as session:
        keyword_row = await session.get(SearchKeyword, search_keyword_id)
        if keyword_row is None:
            return
        now = datetime.now(timezone.utc)
        keyword_row.last_refreshed_at = now
        keyword_row.next_refresh_at = next_refresh_at(now)
        keyword_row.refresh_status = "completed"
        keyword_row.refresh_error = None
        await session.commit()


async def _mark_keyword_refresh_failed(search_keyword_id: UUID, exc: Exception) -> None:
    async with SessionLocal() as session:
        keyword_row = await session.get(SearchKeyword, search_keyword_id)
        if keyword_row is None:
            return
        keyword_row.refresh_status = "failed"
        keyword_row.refresh_error = str(exc)
        await session.commit()


async def _wait_for_keyword_refresh(search_keyword_id: UUID, timeout_seconds: float = 90.0) -> None:
    deadline = asyncio.get_running_loop().time() + timeout_seconds
    while asyncio.get_running_loop().time() < deadline:
        async with SessionLocal() as session:
            row = await session.get(SearchKeyword, search_keyword_id)
            if row is None:
                return
            if row.refresh_status != "running":
                return
        await asyncio.sleep(0.5)


async def _load_keyword_company_pairs(search_keyword_id: UUID, limit: int | None = None) -> list[tuple[SearchKeywordCompany, Company]]:
    async with SessionLocal() as session:
        query = (
            select(SearchKeywordCompany, Company)
            .join(Company, Company.id == SearchKeywordCompany.company_id)
            .where(SearchKeywordCompany.search_keyword_id == search_keyword_id)
            .order_by(SearchKeywordCompany.first_seen_at.asc(), Company.first_seen_at.asc())
        )
        if limit is not None:
            query = query.limit(limit)
        result = await session.execute(query)
        return [(row[0], row[1]) for row in result.all()]


def _company_pair_to_item(link: SearchKeywordCompany, company: Company, keyword: str) -> dict:
    raw_profile = company.raw_profile or {}
    raw_data = {
        **raw_profile,
        "search_title": link.source_title,
        "search_snippet": link.source_snippet,
        "source_query": link.source_query,
        "matched_keywords": [keyword],
    }
    return {
        "company_name": company.company_name,
        "website": company.website,
        "facebook_url": company.facebook_url,
        "linkedin_url": company.linkedin_url,
        "country": company.country,
        "continent": company.continent,
        "source": raw_profile.get("source", "google"),
        "contact_status": "pending",
        "decision_maker_status": "pending",
        "general_contact_status": "pending",
        "contact_name": None,
        "contact_title": None,
        "linkedin_personal_url": None,
        "personal_email": None,
        "work_email": None,
        "phone": None,
        "whatsapp": None,
        "potential_contacts": None,
        "general_emails": [],
        "matched_keywords": [keyword],
        "raw_data": raw_data,
    }


async def _load_keyword_items(search_keyword_id: UUID, keyword: str, limit: int | None) -> list[dict]:
    pairs = await _load_keyword_company_pairs(search_keyword_id, limit=limit)
    return [_company_pair_to_item(link, company, keyword) for link, company in pairs]


async def _get_company_by_domain(domain: str) -> Company | None:
    async with SessionLocal() as session:
        result = await session.execute(select(Company).where(Company.canonical_domain == domain))
        return result.scalar_one_or_none()


async def _upsert_company_profile(profile: dict, *, source_query: str, source_rank: int, source_title: str | None, source_snippet: str | None, search_keyword_id: UUID) -> None:
    now = datetime.now(timezone.utc)
    async with SessionLocal() as session:
        result = await session.execute(select(Company).where(Company.canonical_domain == profile["canonical_domain"]).with_for_update())
        company = result.scalar_one_or_none()
        if company is None:
            company = Company(
                canonical_domain=profile["canonical_domain"],
                company_name=profile.get("company_name"),
                website=profile.get("website"),
                facebook_url=profile.get("facebook_url"),
                linkedin_url=profile.get("linkedin_url"),
                country=profile.get("country"),
                continent=profile.get("continent"),
                raw_profile=profile.get("raw_profile"),
                first_seen_at=now,
                last_seen_at=now,
                last_refreshed_at=now,
            )
            session.add(company)
            await session.flush()
        else:
            company.company_name = profile.get("company_name") or company.company_name
            company.website = profile.get("website") or company.website
            company.facebook_url = profile.get("facebook_url") or company.facebook_url
            company.linkedin_url = profile.get("linkedin_url") or company.linkedin_url
            company.country = profile.get("country") or company.country
            company.continent = profile.get("continent") or company.continent
            company.raw_profile = {**(company.raw_profile or {}), **(profile.get("raw_profile") or {})}
            company.last_seen_at = now
            company.last_refreshed_at = now

        link_result = await session.execute(
            select(SearchKeywordCompany).where(
                SearchKeywordCompany.search_keyword_id == search_keyword_id,
                SearchKeywordCompany.company_id == company.id,
            )
        )
        association = link_result.scalar_one_or_none()
        if association is None:
            association = SearchKeywordCompany(
                search_keyword_id=search_keyword_id,
                company_id=company.id,
                source_query=source_query,
                source_rank=source_rank,
                source_title=source_title,
                source_snippet=source_snippet,
                first_seen_at=now,
                last_seen_at=now,
            )
            session.add(association)
        else:
            association.source_query = source_query or association.source_query
            association.source_rank = source_rank
            association.source_title = source_title or association.source_title
            association.source_snippet = source_snippet or association.source_snippet
            association.last_seen_at = now
        await session.commit()


def _company_is_fresh(company: Company | None) -> bool:
    if company is None or company.last_refreshed_at is None:
        return False
    return (datetime.now(timezone.utc) - company.last_refreshed_at).days < 90


async def _build_lead_item(
    runtime: SearchRuntime,
    serper_item: dict,
    country: str,
    existing_company: Company | None = None,
) -> dict | None:
    url = serper_item.get("link")
    domain = canonical_domain(url)
    if not url or not domain or domain in IGNORED_HOSTS:
        return None
    website = f"https://{domain}"

    if existing_company is not None and _company_is_fresh(existing_company):
        target_country = resolve_country(country)
        if target_country is not None and existing_company.country and existing_company.country != target_country.name_en:
            return None
        raw_profile = existing_company.raw_profile or {}
        return {
            "canonical_domain": existing_company.canonical_domain,
            "company_name": existing_company.company_name,
            "website": existing_company.website or website,
            "facebook_url": existing_company.facebook_url,
            "linkedin_url": existing_company.linkedin_url,
            "country": existing_company.country,
            "continent": existing_company.continent,
            "raw_profile": raw_profile,
            "raw_data": raw_profile,
        }

    async with runtime.scrape_semaphore:
        _, html = await _fetch_homepage(website)
    async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
        name_extractor = CompanyNameExtractor(http_client=client)
        country_detector = CountryDetector(http_client=client)
        social_extractor = SocialLinksExtractor(http_client=client)
        company_name, detection = await asyncio.gather(
            name_extractor.extract(website, serper_result=serper_item, homepage_html=html),
            country_detector.detect(
                website=website,
                target_country=country,
                search_title=serper_item.get("title"),
                search_snippet=serper_item.get("snippet"),
                homepage_html=html,
            ),
        )
        target_country = resolve_country(country)
        if target_country is not None and detection.status != "matched":
            return None
        company_fit = runtime.relevance_classifier.classify(
            website=website,
            company_name=company_name,
            search_title=serper_item.get("title"),
            search_snippet=serper_item.get("snippet"),
            homepage_html=html,
        )
        if not company_fit.is_relevant:
            return None
        social_links = await social_extractor.extract((serper_item.get("knowledgeGraph", {}) or {}).get("title", company_name or domain), domain)
    field_provenance = _merge_field_provenance(
        None,
        {
            "company_fit": _field_source(
                "classification",
                website,
                "industry_relevance_v1",
                company_fit.category,
            ),
            "company_name": _field_source(
                "website",
                website,
                "company_name_extraction_v1",
                serper_item.get("title"),
            ),
            "website": _field_source(
                "search_result",
                serper_item.get("link"),
                "serper_result_link",
                serper_item.get("title"),
            ),
            "facebook_url": social_links.get("facebook_meta"),
            "linkedin_url": social_links.get("linkedin_meta"),
            "country": _field_source(
                "website",
                website,
                "country_detection_v1",
                detection.detected_country_name,
            ),
        },
    )
    raw_profile = {
        "search_title": serper_item.get("title"),
        "search_snippet": serper_item.get("snippet"),
        "name_source": "extraction_v22",
        "source": "google",
        "target_country": target_country.name_en if target_country is not None else country,
        "target_country_code": target_country.code if target_country is not None else None,
        "country_detection": detection.as_dict(),
        "company_fit": company_fit.as_dict(),
        "field_provenance": field_provenance,
    }
    return {
        "canonical_domain": domain,
        "company_name": company_name,
        "website": website,
        "facebook_url": social_links.get("facebook"),
        "linkedin_url": social_links.get("linkedin"),
        "country": detection.detected_country_name,
        "continent": detection.continent,
        "raw_profile": raw_profile,
        "raw_data": raw_profile,
    }


async def _persist_query_frontier(search_keyword_id: UUID, frontier: dict) -> None:
    async with SessionLocal() as session:
        keyword_row = await session.get(SearchKeyword, search_keyword_id)
        if keyword_row is None:
            return
        keyword_row.query_frontier = frontier
        keyword_row.last_requested_at = datetime.now(timezone.utc)
        await session.commit()


async def _refresh_keyword_cache(
    runtime: SearchRuntime,
    *,
    search_keyword_id: UUID,
    child_task_id: UUID | None,
    keyword: str,
    payload: LeadSearchRequest,
    target_count: int | None,
    refresh_one_round: bool = False,
) -> None:
    async with SessionLocal() as session:
        keyword_row = await session.get(SearchKeyword, search_keyword_id)
        frontier = dict(keyword_row.query_frontier or {}) if keyword_row is not None else {}

    pairs = await _load_keyword_company_pairs(search_keyword_id)
    seen_domains = {company.canonical_domain for _, company in pairs}
    pages_limit = _pages_per_query(payload)
    stages = ((1, PRIMARY_QUERY_TEMPLATES), (2, SECONDARY_QUERY_TEMPLATES))

    for stage, _templates in stages:
        query_items = _build_queries(payload, keyword=keyword, stage=stage)
        progress_in_round = True
        first_round = True
        while progress_in_round:
            progress_in_round = False
            for query_info in query_items:
                query_key = str(query_info["query_key"])
                query_state = frontier.get(query_key, {"next_page": 1, "exhausted": False})
                if query_state.get("exhausted"):
                    continue
                page = int(query_state.get("next_page", 1))
                if page > pages_limit:
                    query_state["exhausted"] = True
                    frontier[query_key] = query_state
                    continue
                progress_in_round = True
                async with runtime.search_semaphore:
                    data = await _safe_serper_search(
                        runtime.serper_client,
                        query=str(query_info["query"]),
                        hl=str(query_info["language"]),
                        gl=str(query_info["gl"]),
                        num=QUERY_PAGE_SIZE,
                        page=page,
                    )
                if child_task_id is not None:
                    await _mutate_task(child_task_id, phase="querying", processed_search_requests_inc=1)
                organic = data.get("organic", [])
                query_state["next_page"] = page + 1
                if len(organic) < QUERY_PAGE_SIZE:
                    query_state["exhausted"] = True
                frontier[query_key] = query_state
                await _persist_query_frontier(search_keyword_id, frontier)

                for rank, item in enumerate(organic, start=1):
                    if child_task_id is not None:
                        await _mutate_task(child_task_id, processed_candidates_inc=1, phase="building_leads")
                    domain = canonical_domain(item.get("link"))
                    if not domain or domain in seen_domains or domain in IGNORED_HOSTS:
                        continue
                    existing_company = await _get_company_by_domain(domain)
                    profile = await _build_lead_item(
                        runtime,
                        item,
                        str(query_info["country"]),
                        existing_company=existing_company,
                    )
                    if profile is None:
                        continue
                    await _upsert_company_profile(
                        profile,
                        source_query=str(query_info["query"]),
                        source_rank=rank,
                        source_title=item.get("title"),
                        source_snippet=item.get("snippet"),
                        search_keyword_id=search_keyword_id,
                    )
                    seen_domains.add(profile["canonical_domain"])
                    if child_task_id is not None:
                        await _mutate_task(child_task_id, confirmed_leads=len(seen_domains))
                    if target_count is not None and len(seen_domains) >= target_count:
                        break
                if target_count is not None and len(seen_domains) >= target_count:
                    break
            if refresh_one_round and first_round:
                break
            first_round = False
            if target_count is not None and len(seen_domains) >= target_count:
                break
        if target_count is not None and len(seen_domains) >= target_count:
            break

    await _complete_keyword_refresh(search_keyword_id)


async def _finalize_child_task(child_task_id: UUID, *, keyword: str, items: list[dict], cache_hit: bool) -> None:
    parent_task_id: UUID | None = None
    async with SessionLocal() as session:
        result = await session.execute(select(Task).where(Task.id == child_task_id).with_for_update())
        task = result.scalar_one_or_none()
        if task is None:
            return
        params = {**(task.params or {}), "keyword": keyword, "cache_hit": cache_hit}
        task.params = params
        task.status = "completed"
        task.phase = "completed"
        task.progress = 100
        task.completed = max(task.processed_candidates, len(items))
        task.confirmed_leads = len(items)
        task.estimated_remaining_seconds = 0
        task.updated_at = datetime.now(timezone.utc)
        parent_task_id = task.parent_task_id
        await session.commit()
    if parent_task_id is not None:
        await _recompute_root_task(parent_task_id)


async def _run_keyword_search_task(
    runtime: SearchRuntime,
    *,
    child_task_id: UUID,
    root_task_id: UUID,
    keyword: str,
    payload: LeadSearchRequest,
) -> dict:
    await _mutate_task(child_task_id, status="running", phase="checking_cache")
    if payload.mode == "demo":
        items = _build_demo_keyword_items(payload, keyword)
        limit = _keyword_limit(payload)
        if limit is not None:
            items = items[:limit]
        await _finalize_child_task(child_task_id, keyword=keyword, items=items, cache_hit=False)
        return {"keyword": keyword, "items": items, "cache_hit": False}

    keyword_row = await _get_or_create_search_keyword(keyword, payload)
    limit = _keyword_limit(payload)
    cached_items = await _load_keyword_items(keyword_row.id, keyword, limit=limit)
    if limit is None:
        if cached_items:
            await _finalize_child_task(child_task_id, keyword=keyword, items=cached_items, cache_hit=True)
            return {"keyword": keyword, "items": cached_items, "cache_hit": True}
    elif len(cached_items) >= limit:
        await _finalize_child_task(child_task_id, keyword=keyword, items=cached_items[:limit], cache_hit=True)
        return {"keyword": keyword, "items": cached_items[:limit], "cache_hit": True}

    claimed = await _claim_keyword_refresh(keyword_row.id)
    if claimed:
        try:
            fill_target = limit
            if fill_target is None:
                fill_target = max(10, len(payload.countries or [""]) * len(payload.languages or ["en"]))
            await _refresh_keyword_cache(
                runtime,
                search_keyword_id=keyword_row.id,
                child_task_id=child_task_id,
                keyword=keyword,
                payload=payload,
                target_count=fill_target,
            )
        except Exception as exc:
            await _mark_keyword_refresh_failed(keyword_row.id, exc)
            await _mutate_task(child_task_id, status="failed", phase="failed", error=str(exc), estimated_remaining_seconds=0)
            raise
    else:
        await _mutate_task(child_task_id, phase="waiting_cache")
        await _wait_for_keyword_refresh(keyword_row.id)

    refreshed_items = await _load_keyword_items(keyword_row.id, keyword, limit=limit)
    await _finalize_child_task(child_task_id, keyword=keyword, items=refreshed_items, cache_hit=False)
    return {"keyword": keyword, "items": refreshed_items, "cache_hit": False}


def _merge_keyword_results(keyword_results: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    ordered_domains: list[str] = []
    for keyword_result in keyword_results:
        keyword = keyword_result["keyword"]
        for item in keyword_result["items"]:
            domain = canonical_domain(item.get("website")) or item.get("company_name") or f"row-{len(ordered_domains)}"
            existing = merged.get(domain)
            if existing is None:
                payload = {**item}
                payload["matched_keywords"] = list(item.get("matched_keywords") or [keyword])
                payload["raw_data"] = {**(item.get("raw_data") or {}), "matched_keywords": list(payload["matched_keywords"])}
                merged[domain] = payload
                ordered_domains.append(domain)
                continue
            matched_keywords = list(existing.get("matched_keywords") or [])
            if keyword not in matched_keywords:
                matched_keywords.append(keyword)
            existing["matched_keywords"] = matched_keywords
            raw_data = {**(existing.get("raw_data") or {})}
            raw_data["matched_keywords"] = matched_keywords
            keyword_sources = list(raw_data.get("keyword_sources") or [])
            keyword_sources.append(
                {
                    "keyword": keyword,
                    "search_title": (item.get("raw_data") or {}).get("search_title"),
                    "search_snippet": (item.get("raw_data") or {}).get("search_snippet"),
                    "source_query": (item.get("raw_data") or {}).get("source_query"),
                }
            )
            raw_data["keyword_sources"] = keyword_sources
            existing["raw_data"] = raw_data
    return [merged[domain] for domain in ordered_domains]


async def _persist_root_results(task_id: UUID, results: list[dict], *, error: str | None = None) -> None:
    async with SessionLocal() as session:
        result = await session.execute(select(Task).where(Task.id == task_id).with_for_update())
        task = result.scalar_one_or_none()
        if task is None:
            return
        await session.execute(delete(Lead).where(Lead.task_id == task_id))
        now = datetime.now(timezone.utc)
        session.add_all([Lead(task_id=task_id, created_at=now, **item) for item in results])
        task.status = "completed" if results or not error else "failed"
        task.phase = "completed" if results or not error else "failed"
        task.progress = 100 if results or not error else task.progress
        task.completed = max(task.processed_candidates, len(results))
        task.confirmed_leads = len(results)
        task.error = error
        task.estimated_remaining_seconds = 0
        task.updated_at = now
        await session.commit()


async def _mark_root_failed(task_id: UUID, exc: Exception) -> None:
    async with SessionLocal() as session:
        result = await session.execute(select(Task).where(Task.id == task_id).with_for_update())
        task = result.scalar_one_or_none()
        if task is None:
            return
        task.status = "failed"
        task.phase = "failed"
        task.error = str(exc)
        task.estimated_remaining_seconds = 0
        task.updated_at = datetime.now(timezone.utc)
        await session.commit()


async def _run_root_search_task(task_id: UUID, payload: LeadSearchRequest) -> None:
    runtime = SearchRuntime()
    try:
        await _mutate_task(task_id, status="running", phase="querying")
        async with SessionLocal() as session:
            result = await session.execute(
                select(Task).where(Task.parent_task_id == task_id, Task.type == KEYWORD_TASK_TYPE).order_by(Task.created_at.asc())
            )
            children = result.scalars().all()
        child_results = await asyncio.gather(
            *[
                _run_keyword_search_task(runtime, child_task_id=child.id, root_task_id=task_id, keyword=(child.params or {}).get("keyword", ""), payload=payload)
                for child in children
            ],
            return_exceptions=True,
        )
        success_results: list[dict] = []
        errors: list[str] = []
        for child_result in child_results:
            if isinstance(child_result, Exception):
                errors.append(str(child_result))
                continue
            success_results.append(child_result)
        if not success_results and errors:
            raise RuntimeError("；".join(errors))
        merged_results = _merge_keyword_results(success_results)
        await _persist_root_results(task_id, merged_results, error="；".join(errors) if errors else None)
    except Exception as exc:
        await _mark_root_failed(task_id, exc)


def _style_workbook(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    widths = {"A": 24, "B": 30, "C": 26, "D": 28, "E": 22, "F": 18, "G": 16, "H": 22, "I": 22, "J": 24, "K": 28, "L": 20, "M": 20, "N": 30}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _lead_field_value(lead: Lead, field_key: str) -> str | None:
    if field_key == "company_fit":
        raw_data = lead.raw_data or {}
        company_fit = raw_data.get("company_fit") or {}
        category = company_fit.get("category")
        return str(category) if category else None
    value = getattr(lead, field_key, None)
    if value is None and field_key == "potential_contacts":
        contacts = (lead.potential_contacts or {}).get("items", [])
        return "\n".join(str(item) for item in contacts if item)
    if isinstance(value, list):
        return "\n".join(str(item) for item in value if item)
    if isinstance(value, dict):
        return "\n".join(str(item) for item in value.get("items", []) if item)
    return str(value) if value not in {None, ""} else None


def _ensure_reviewable_field(field_key: str) -> str:
    normalized = str(field_key or "").strip()
    if normalized not in REVIEWABLE_FIELD_KEYS:
        raise HTTPException(status_code=400, detail=f"Unsupported review field: {normalized}")
    return normalized


@router.post("/search", response_model=TaskCreateResponse)
async def create_lead_search(payload: LeadSearchRequest, db: AsyncSession = Depends(get_db)) -> TaskCreateResponse:
    runtime_plan = _estimate_search_runtime(payload)
    now = datetime.now(timezone.utc)
    root_task = Task(
        type=ROOT_TASK_TYPE,
        status="pending",
        progress=0,
        total=runtime_plan["planned_candidate_budget"],
        completed=0,
        target_count=payload.target_count,
        confirmed_leads=0,
        stopped_early=False,
        params={
            "keywords": payload.keywords,
            "product_name": payload.keywords[0],
            "continents": payload.continents,
            "countries": payload.countries,
            "languages": payload.languages,
            "mode": payload.mode,
        },
        estimated_total_seconds=runtime_plan["estimated_total_seconds"],
        estimated_remaining_seconds=runtime_plan["estimated_total_seconds"],
        phase="queued",
        processed_search_requests=0,
        planned_search_requests=runtime_plan["planned_search_requests"],
        processed_candidates=0,
        planned_candidate_budget=runtime_plan["planned_candidate_budget"],
        created_at=now,
        updated_at=now,
    )
    db.add(root_task)
    await db.flush()

    for keyword in payload.keywords:
        keyword_plan = _estimate_keyword_runtime(payload, keyword)
        db.add(
            Task(
                parent_task_id=root_task.id,
                type=KEYWORD_TASK_TYPE,
                status="pending",
                progress=0,
                total=keyword_plan["planned_candidate_budget"],
                completed=0,
                target_count=payload.target_count,
                confirmed_leads=0,
                stopped_early=False,
                params={
                    "keyword": keyword,
                    "keywords": [keyword],
                    "product_name": keyword,
                    "continents": payload.continents,
                    "countries": payload.countries,
                    "languages": payload.languages,
                    "mode": payload.mode,
                    "cache_hit": False,
                },
                estimated_total_seconds=keyword_plan["estimated_total_seconds"],
                estimated_remaining_seconds=keyword_plan["estimated_total_seconds"],
                phase="queued",
                processed_search_requests=0,
                planned_search_requests=keyword_plan["planned_search_requests"],
                processed_candidates=0,
                planned_candidate_budget=keyword_plan["planned_candidate_budget"],
                created_at=now,
                updated_at=now,
            )
        )

    await db.commit()
    await db.refresh(root_task)

    await cleanup_old_root_tasks(db)
    await db.commit()

    asyncio.create_task(_run_root_search_task(root_task.id, payload.model_copy(deep=True)))
    return TaskCreateResponse(task_id=root_task.id)


@router.get("", response_model=LeadListResponse)
async def list_leads(
    task_id: UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
) -> LeadListResponse:
    return await get_task_leads(db, task_id, page=page, page_size=page_size)


@router.get("/export")
async def export_leads(task_id: UUID, format: str = "xlsx", include_contacts: bool = False, db: AsyncSession = Depends(get_db)):
    lead_page = await get_task_leads(db, task_id, page=1, page_size=200)
    leads = lead_page.items
    while len(leads) < lead_page.total:
        next_page = (len(leads) // 200) + 1
        next_items = await get_task_leads(db, task_id, page=next_page, page_size=200)
        leads.extend(next_items.items)

    headers = ["company_name", "website", "facebook_url", "linkedin_url", "matched_keywords", "detected_country", "status", "name_source"]
    if include_contacts:
        headers += ["contact_name", "contact_title", "personal_email", "work_email", "phone", "whatsapp", "potential_contacts"]
    if format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for lead in leads:
            row = [
                lead.company_name or "",
                lead.website or "",
                lead.facebook_url or "",
                lead.linkedin_url or "",
                ", ".join(lead.matched_keywords or []),
                lead.country or "",
                lead.contact_status,
                (lead.raw_data or {}).get("name_source", ""),
            ]
            if include_contacts:
                row += [lead.contact_name or "", lead.contact_title or "", lead.personal_email or "", lead.work_email or "", lead.phone or "", lead.whatsapp or "", "; ".join((lead.potential_contacts or {}).get("items", []))]
            writer.writerow(row)
        return Response(content=buffer.getvalue(), media_type="text/csv")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lead Results"
    pretty_headers = ["Company Name", "Website", "Facebook URL", "LinkedIn URL", "Matched Keywords", "Detected Country", "Status", "Name Source"]
    if include_contacts:
        pretty_headers += ["Contact Name", "Contact Title", "Personal Email", "Work Email", "Phone", "WhatsApp", "Potential Contacts"]
    sheet.append(pretty_headers)
    for lead in leads:
        row = [
            lead.company_name,
            lead.website,
            lead.facebook_url,
            lead.linkedin_url,
            ", ".join(lead.matched_keywords or []),
            lead.country,
            lead.contact_status,
            (lead.raw_data or {}).get("name_source"),
        ]
        if include_contacts:
            row += [lead.contact_name, lead.contact_title, lead.personal_email, lead.work_email, lead.phone, lead.whatsapp, "\n".join((lead.potential_contacts or {}).get("items", []))]
        sheet.append(row)
    _style_workbook(sheet)
    data = BytesIO()
    workbook.save(data)
    data.seek(0)
    return StreamingResponse(data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.put("/{lead_id}/reviews/{field_key}", response_model=LeadReviewAnnotation)
async def upsert_lead_review(
    lead_id: UUID,
    field_key: str,
    payload: LeadReviewUpsertRequest,
    db: AsyncSession = Depends(get_db),
) -> LeadReviewAnnotation:
    normalized_field_key = _ensure_reviewable_field(field_key)
    lead = await db.get(Lead, lead_id)
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")
    result = await db.execute(
        select(LeadReview).where(
            LeadReview.lead_id == lead_id,
            LeadReview.field_key == normalized_field_key,
        )
    )
    review = result.scalar_one_or_none()
    now = datetime.now(timezone.utc)
    if review is None:
        review = LeadReview(
            lead_id=lead_id,
            field_key=normalized_field_key,
            verdict=payload.verdict,
            source_path=payload.source_path,
            note=payload.note,
            created_at=now,
            updated_at=now,
        )
        db.add(review)
    else:
        review.verdict = payload.verdict
        review.source_path = payload.source_path
        review.note = payload.note
        review.updated_at = now
    await db.commit()
    await db.refresh(review)
    return _review_annotation(review)


@router.delete("/{lead_id}/reviews/{field_key}")
async def delete_lead_review(lead_id: UUID, field_key: str, db: AsyncSession = Depends(get_db)) -> dict:
    normalized_field_key = _ensure_reviewable_field(field_key)
    result = await db.execute(
        select(LeadReview).where(
            LeadReview.lead_id == lead_id,
            LeadReview.field_key == normalized_field_key,
        )
    )
    review = result.scalar_one_or_none()
    if review is not None:
        await db.delete(review)
        await db.commit()
    return {"deleted": True}


@router.delete("")
async def delete_leads(task_id: UUID, db: AsyncSession = Depends(get_db)) -> dict:
    task = await get_root_task(db, task_id)
    if task is not None:
        await db.delete(task)
        await db.commit()
    return {"deleted": True}
